# -*- coding: utf-8 -*-
from openpyxl import Workbook


class Create_workbook:
    def __init__(self,sheet_name):

        self.write_wb = Workbook()
        # 이름이 있는 시트를 생성
        # self.write_ws = self.write_wb.create_sheet(sheet_name)

        #생성된 시트를 선택
        self.write_ws = self.write_wb.active
        self.write_ws.title = sheet_name

    def input(self,index,value):
        self.write_ws[index]=value
        # write_ws['A1'] = '숫자'

    def input_line(self,list):
        self.write_ws.append(list)
        # write_ws.append([1, 2, 3])

    def input_cell(self,row_idx,column_idx,value):
        self.write_ws.cell(row=row_idx,column=column_idx,value=value)
        # write_ws.cell(5, 5, '5행5열')


    def save_file(self,excle_path):
        try:
            self.write_wb.save(excle_path)
        except:
            print("파일 저장에 실패 했습니다.")
        # write_wb.save('/Users/Jamong/Desktop/숫자.xlsx')

# # Sheet1에다 입력
# write_ws = write_wb.active
# write_ws['A1'] = '숫자'
#
# # 행 단위로 추가
# write_ws.append([1, 2, 3])
#
# # 셀 단위로 추가
# write_ws.cell(5, 5, '5행5열')
# write_wb.save('/Users/Jamong/Desktop/숫자.xlsx')